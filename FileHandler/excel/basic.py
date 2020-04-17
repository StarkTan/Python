"""
创建修改 Excel 文件
1. 创建文件
2. 创建表单
3. 向表格中写值
4. 设置表格的风格
5. 修改行和列的高宽
6. 表格合并
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

wb = openpyxl.Workbook() # 创建文件
# wb = openpyxl.load_workbook('filetest.xlsx') # 已有的模板创建文件

print("Sheet 操作")
cur_sheet = wb.active  # 获取默认打开的Sheet
print(cur_sheet.title)

wb.remove(cur_sheet)  # 删除Sheet

wb.create_sheet("Sheet1", 1)
wb.create_sheet("Sheet2", 3)

sheet_names = wb.get_sheet_names()  # 获取sheet列表
print(sheet_names)
sheet = wb.get_sheet_by_name("Sheet1")  # 根据名称找到对应Sheet

print("Cell 写值，读值")
sheet['F6'] = 'Stark'
print(sheet['F6'].value)

print("Style 操作")
sheet['F6'].font = Font(size=14, italic=True)  # 字体设置
sheet['F6'].fill = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='000000')  # 背景填充
sheet['F6'].border = Border(left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'))  # 边框样式 源码还有更多
sheet['F6'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)  # 对其方式  wrap_text（自动换行）

print("Cell 写公式")
sheet['A1'] = 2
sheet['A2'] = 5
sheet['A4'] = '=SUM(A1:A3)'

print("修改行列的高宽")
sheet.row_dimensions[3].height = 65
sheet.column_dimensions['F'].width = 25

print("单元格合并")
sheet['B2'] = "stark"
sheet['B3'] = "xxx"
sheet.merge_cells('B2:D3')  # 合并单元格，只会保留最左上角的内容
print(sheet['B2'].value, sheet['B3'].value)
sheet.unmerge_cells('B2:D3')  # 取消单元格合并
print(sheet['B2'].value, sheet['B3'].value)

wb.save('test.xlsx')  # 保存文件
wb.close()  # 保存文件后需要关闭，否则文件无法读取
